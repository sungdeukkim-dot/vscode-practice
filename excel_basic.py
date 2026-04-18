from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ================================
# 1. 워크북 생성
# ================================
wb = Workbook()
ws = wb.active
ws.title = "포트폴리오"

# ================================
# 2. 헤더 작성
# ================================
headers = ["No.", "회사명", "성장률(%)", "매출액(억)", "투자단계", "판정"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)

    # 헤더 스타일
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", fgColor="2E4057")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ================================
# 3. 데이터 입력
# ================================
data = [
    (1, "SemiFive",  85,  120, "Series B", "Good"),
    (2, "CYTUR",     42,   30, "Series A", "Caution"),
    (3, "Mearitt",  110,   80, "Series B", "Excellent"),
    (4, "인탑스",    75,   55, "Series A", "Good"),
    (5, "Olive",     60,   45, "Series A", "Good"),
]

for row_data in data:
    ws.append(row_data)

# ================================
# 4. 데이터 행 스타일
# ================================
verdict_colors = {
    "Excellent": "C8F7C5",  # 연초록
    "Good":      "D6EAF8",  # 연파랑
    "Caution":   "FDEBD0",  # 연주황
    "Reject":    "FADBD8",  # 연빨강
}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    verdict = row[5].value
    color   = verdict_colors.get(verdict, "FFFFFF")

    for cell in row:
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")

# ================================
# 5. 열 너비 자동 조정
# ================================
col_widths = [6, 16, 12, 12, 12, 12]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ================================
# 6. 행 높이 설정
# ================================
ws.row_dimensions[1].height = 25
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 20

# ================================
# 7. 저장
# ================================
wb.save("portfolio_report.xlsx")
print("Excel 보고서 생성 완료: portfolio_report.xlsx")