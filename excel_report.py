from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import json
import os


# ================================
# 투자 판정 함수
# ================================
def evaluate(growth, revenue):
    g = 3 if growth >= 100 else 2 if growth >= 50 else 1 if growth >= 0 else 0
    r = 3 if revenue >= 100 else 2 if revenue >= 50 else 1 if revenue >= 10 else 0
    total = g + r
    if total >= 5:
        return "Excellent"
    elif total >= 3:
        return "Good"
    elif total >= 1:
        return "Caution"
    else:
        return "Reject"


# ================================
# 테두리 스타일
# ================================
def thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ================================
# 보고서 생성
# ================================
def create_report(data):
    wb = Workbook()

    # ================================
    # 시트 1. 포트폴리오 현황
    # ================================
    ws1 = wb.active
    ws1.title = "포트폴리오 현황"

    # 타이틀
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value     = f"투자 포트폴리오 분석 보고서  |  생성일: {datetime.now().strftime('%Y-%m-%d')}"
    title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="1A1A2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35

    # 헤더
    headers = ["No.", "회사명", "성장률(%)", "매출액(억)", "투자단계", "종합점수", "판정"]
    for col, h in enumerate(headers, 1):
        cell            = ws1.cell(row=2, column=col, value=h)
        cell.font       = Font(bold=True, color="FFFFFF", size=10)
        cell.fill       = PatternFill("solid", fgColor="16213E")
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = thin_border()
    ws1.row_dimensions[2].height = 22

    # 데이터
    verdict_colors = {
        "Excellent": "C8F7C5",
        "Good":      "D6EAF8",
        "Caution":   "FDEBD0",
        "Reject":    "FADBD8",
    }

    for i, item in enumerate(data, 1):
        verdict = evaluate(item["growth"], item["revenue"])
        score   = round(item["growth"] * 0.6 + item["revenue"] * 0.4, 1)
        color   = verdict_colors.get(verdict, "FFFFFF")
        row     = i + 2

        values = [i, item["company"], item["growth"],
                  item["revenue"], item["stage"], score, verdict]

        for col, val in enumerate(values, 1):
            cell            = ws1.cell(row=row, column=col, value=val)
            cell.fill       = PatternFill("solid", fgColor=color)
            cell.alignment  = Alignment(horizontal="center", vertical="center")
            cell.border     = thin_border()
        ws1.row_dimensions[row].height = 20

    # 열 너비
    widths = [6, 16, 12, 12, 14, 12, 12]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ================================
    # 시트 2. 통계 요약
    # ================================
    ws2 = wb.create_sheet("통계 요약")

    ws2.merge_cells("A1:C1")
    ws2["A1"].value     = "통계 요약"
    ws2["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill      = PatternFill("solid", fgColor="1A1A2E")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 30

    growths  = [d["growth"]  for d in data]
    revenues = [d["revenue"] for d in data]

    stats = [
        ("총 검토 회사 수",  f"{len(data)}개"),
        ("평균 성장률",      f"{sum(growths)/len(growths):.1f}%"),
        ("최고 성장률",      f"{max(growths)}%"),
        ("최저 성장률",      f"{min(growths)}%"),
        ("평균 매출액",      f"{sum(revenues)/len(revenues):.1f}억"),
        ("최고 매출액",      f"{max(revenues)}억"),
        ("Excellent 수",    f"{sum(1 for d in data if evaluate(d['growth'], d['revenue']) == 'Excellent')}개"),
        ("Good 수",         f"{sum(1 for d in data if evaluate(d['growth'], d['revenue']) == 'Good')}개"),
        ("Caution/Reject",  f"{sum(1 for d in data if evaluate(d['growth'], d['revenue']) in ['Caution','Reject'])}개"),
    ]

    for row, (label, value) in enumerate(stats, 2):
        ws2.cell(row=row, column=1, value=label).font  = Font(bold=True)
        ws2.cell(row=row, column=1).fill               = PatternFill("solid", fgColor="EBF5FB")
        ws2.cell(row=row, column=1).border             = thin_border()
        ws2.cell(row=row, column=2, value=value).border = thin_border()
        ws2.cell(row=row, column=2).alignment           = Alignment(horizontal="center")
        ws2.row_dimensions[row].height = 20

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14

    # ================================
    # 시트 3. 차트
    # ================================
    ws3 = wb.create_sheet("차트")

    ws3.append(["회사명", "성장률", "매출액"])
    for item in data:
        ws3.append([item["company"], item["growth"], item["revenue"]])

    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Portfolio Analysis"
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Company"

    data_ref   = Reference(ws3, min_col=2, max_col=3, min_row=1, max_row=len(data)+1)
    cats       = Reference(ws3, min_col=1, min_row=2, max_row=len(data)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.width  = 20
    chart.height = 14

    ws3.add_chart(chart, "E2")

    # ================================
    # 저장
    # ================================
    filename = f"investment_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    print(f"\nExcel 보고서 생성 완료: {filename}\n")
    return filename


# ================================
# 실행
# ================================
DATA_FILE = "investment_data.json"

if os.path.exists(DATA_FILE):
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
else:
    # 샘플 데이터 (investment_data.json 없을 때)
    data = [
        {"company": "SemiFive",  "growth": 85,  "revenue": 120, "stage": "Series B"},
        {"company": "CYTUR",     "growth": 42,  "revenue": 30,  "stage": "Series A"},
        {"company": "Mearitt",   "growth": 110, "revenue": 80,  "stage": "Series B"},
        {"company": "인탑스",    "growth": 75,  "revenue": 55,  "stage": "Series A"},
        {"company": "Olive",     "growth": 60,  "revenue": 45,  "stage": "Series A"},
    ]

create_report(data)