import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def get_stock_info(code):
    url     = f"https://finance.naver.com/item/main.naver?code={code}"
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        res  = requests.get(url, headers=headers, timeout=5)
        soup = BeautifulSoup(res.text, "html.parser")

        name_tag  = soup.select_one(".wrap_company h2 a")
        price_tag = soup.select_one(".no_today .blind")
        change_tag = soup.select_one(".no_exday .blind")

        name   = name_tag.get_text(strip=True)  if name_tag  else "N/A"
        price  = price_tag.get_text(strip=True)  if price_tag  else "N/A"
        change = change_tag.get_text(strip=True) if change_tag else "N/A"

        market_cap = "N/A"
        for tr in soup.select(".tb_type1 tr"):
            th = tr.select_one("th")
            if th and "시가총액" in th.get_text():
                td = tr.select_one("td")
                if td:
                    market_cap = td.get_text(strip=True)
                break

        return {
            "code": code, "name": name,
            "price": price, "change": change,
            "market_cap": market_cap,
        }
    except Exception as e:
        return {"code": code, "name": "오류", "price": "-",
                "change": "-", "market_cap": "-"}


def create_stock_excel(stocks):
    wb = Workbook()
    ws = wb.active
    ws.title = "시세 현황"

    # 타이틀
    ws.merge_cells("A1:F1")
    ws["A1"].value     = f"주요 종목 시세  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A1A2E")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # 헤더
    headers = ["종목코드", "종목명", "현재가", "전일대비", "시가총액", "수집시각"]
    thin    = Side(style="thin", color="CCCCCC")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=2, column=col, value=h)
        cell.font       = Font(bold=True, color="FFFFFF", size=10)
        cell.fill       = PatternFill("solid", fgColor="16213E")
        cell.alignment  = Alignment(horizontal="center")
        cell.border     = border
    ws.row_dimensions[2].height = 22

    # 데이터
    now = datetime.now().strftime("%H:%M:%S")
    for i, stock in enumerate(stocks, 3):
        row_color = "F2F3F4" if i % 2 == 0 else "FFFFFF"
        values    = [stock["code"], stock["name"], stock["price"],
                     stock["change"], stock["market_cap"], now]
        for col, val in enumerate(values, 1):
            cell            = ws.cell(row=i, column=col, value=val)
            cell.fill       = PatternFill("solid", fgColor=row_color)
            cell.alignment  = Alignment(horizontal="center")
            cell.border     = border
        ws.row_dimensions[i].height = 20

    # 열 너비
    for i, w in enumerate([12, 16, 12, 12, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    print(f"\nExcel 저장 완료: {filename}\n")


# ================================
# 실행
# ================================
watchlist = [
    "005930",  # 삼성전자
    "000660",  # SK하이닉스
    "035420",  # NAVER
    "051910",  # LG화학
    "006400",  # 삼성SDI
]

print("데이터 수집 중...\n")
stocks = [get_stock_info(code) for code in watchlist]

for s in stocks:
    print(f"{s['name']}: {s['price']} ({s['change']})")

create_stock_excel(stocks)